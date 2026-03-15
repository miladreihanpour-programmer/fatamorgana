from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


THIN_SIDE = Side(style="thin", color="999999")
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
METRIC_HEADERS = {"ESAURITO", "MANTENIMENTO", "ORDINE"}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Format an Excel workbook for PDF conversion.")
    parser.add_argument("excel_path", help="Path to the source workbook.")
    parser.add_argument("output_path", nargs="?", default=None, help="Optional output workbook path.")
    parser.add_argument("--sheet", default=None, help="Optional sheet name. Defaults to the first sheet.")
    parser.add_argument(
        "--only-sheet",
        action="store_true",
        help="Keep only the selected sheet in output workbook.",
    )
    return parser.parse_args()


def used_range(ws: Worksheet) -> tuple[int, int]:
    max_row = 1
    max_col = 1
    for row in ws.iter_rows():
        for cell in row:
            if cell.value not in (None, ""):
                max_row = max(max_row, cell.row)
                max_col = max(max_col, cell.column)
    return max_row, max_col


def format_value(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def detect_group_start_columns(ws: Worksheet, max_col: int) -> set[int]:
    starts: set[int] = set()
    for col_idx in range(1, max_col + 1):
        header = format_value(ws.cell(row=1, column=col_idx).value).upper()
        if header and header not in METRIC_HEADERS:
            starts.add(col_idx)
    if not starts:
        starts.add(1)
    return starts


def apply_layout(ws: Worksheet) -> None:
    max_row, max_col = used_range(ws)

    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_setup.scale = None
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.sheet_properties.pageSetUpPr.autoPageBreaks = False
    ws.page_margins.left = 0.05
    ws.page_margins.right = 0.05
    ws.page_margins.top = 0.12
    ws.page_margins.bottom = 0.12
    ws.page_margins.header = 0.05
    ws.page_margins.footer = 0.05
    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = False
    ws.print_title_rows = "$1:$1"
    ws.freeze_panes = "A2"

    ws.print_area = f"A1:{get_column_letter(max_col)}{max_row}"

    for row_idx in range(1, max_row + 1):
        ws.row_dimensions[row_idx].height = 20

    group_start_cols = detect_group_start_columns(ws, max_col)
    group_starts_sorted = sorted(group_start_cols)
    header_fill_palette = ["F3E6D6", "E6F0E8", "E7ECF7", "F8E4E4"]

    from openpyxl.styles import PatternFill

    # Adaptive widths based on text lengths, then scaled to fill printable A4 width.
    raw_widths: dict[int, float] = {}
    for col_idx in range(1, max_col + 1):
        values = [format_value(ws.cell(row=row_idx, column=col_idx).value) for row_idx in range(1, max_row + 1)]
        max_len = max((len(v) for v in values), default=0)

        if col_idx in group_start_cols:
            # Flavor columns
            width = min(max((max_len * 0.80) + 3.0, 15.0), 34.0)
        else:
            # Numeric/status columns
            width = min(max((max_len * 0.55) + 1.2, 4.0), 10.5)

        raw_widths[col_idx] = width

    total_width = sum(raw_widths.values()) or 1.0
    target_total_width = 182.0
    scale = target_total_width / total_width

    for col_idx in range(1, max_col + 1):
        column_letter = get_column_letter(col_idx)
        width = raw_widths[col_idx] * scale
        if col_idx in group_start_cols:
            width = min(max(width, 15.0), 36.0)
        else:
            width = min(max(width, 3.8), 9.0)
        ws.column_dimensions[column_letter].width = width

    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.border = THIN_BORDER
            cell.font = Font(name="Calibri", size=10, bold=(cell.row == 1))
            if cell.column in group_start_cols:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False, shrink_to_fit=False)
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False, shrink_to_fit=False)

            if cell.row == 1 and cell.column in group_start_cols:
                group_idx = group_starts_sorted.index(cell.column)
                fill_color = header_fill_palette[group_idx % len(header_fill_palette)]
                next_start = group_starts_sorted[group_idx + 1] if group_idx + 1 < len(group_starts_sorted) else (max_col + 1)
                for group_col in range(cell.column, next_start):
                    ws.cell(row=1, column=group_col).fill = PatternFill(fill_type="solid", fgColor=fill_color)


def main() -> None:
    args = parse_args()
    excel_path = Path(args.excel_path)
    if not excel_path.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")

    output_path = Path(args.output_path) if args.output_path else excel_path
    workbook = load_workbook(excel_path)
    sheet_name = args.sheet or workbook.sheetnames[0]
    if args.only_sheet:
        for name in list(workbook.sheetnames):
            if name != sheet_name:
                workbook.remove(workbook[name])
    ws = workbook[sheet_name]
    apply_layout(ws)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    print(f"Formatted workbook saved to: {output_path}")


if __name__ == "__main__":
    main()