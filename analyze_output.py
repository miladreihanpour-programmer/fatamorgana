#!/usr/bin/env python3
import openpyxl
import sys

def check_file(filepath):
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active
        
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:  # if first column has value
                rows.append(row)
        
        print(f"✓ {filepath}")
        print(f"  Total rows: {len(rows)}")
        
        if len(rows) > 0:
            print(f"  Columns: {ws[1]}")
            print(f"\n  First 5 rows:")
            for i, row in enumerate(rows[:5]):
                print(f"    {i+1}. {str(row[0])[:30]:<30} {row[1:5]}")
            
            if len(rows) > 5:
                print(f"\n  Last 5 rows:")
                for i, row in enumerate(rows[-5:], start=len(rows)-4):
                    print(f"    {i}. {str(row[0])[:30]:<30} {row[1:5]}")
        
        return len(rows)
    except Exception as e:
        print(f"✗ Error reading {filepath}: {e}")
        return 0

print("=" * 70)
print("📊 ANALYZING OUTPUT FILES")
print("=" * 70)

print("\n1. MANTENIMENTO (Current Stock)")
man_count = check_file('d:\\fata\\output\\shocapp_mantenimento.xlsx')

print("\n2. ESAURITO (Sold Last 7 Days)")
esau_count = check_file('d:\\fata\\output\\shocapp_esaurito.xlsx')

print("\n3. DA ORDINARE (Final Decisions)")
ord_count = check_file('d:\\fata\\output\\shocapp_da_ordinare.xlsx')

print("\n" + "=" * 70)
print("📈 SUMMARY")
print("=" * 70)
print(f"Stock entries: {man_count}")
print(f"Sales entries: {esau_count}")
print(f"Order decisions: {ord_count}")
print()

if ord_count == 68:
    print("⚠️  ISSUE FOUND: 68 decision rows (bug not fixed)")
    print("   This means code changes didn't execute or extraction wasn't re-run")
elif ord_count > 30:
    print("⚠️  ISSUE: Still too many decisions")
else:
    print("✅ OK: Reasonable number of decisions")
